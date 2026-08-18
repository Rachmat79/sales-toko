# -*- coding: utf-8 -*-
"""
SALES DISPLAY ANALYZER
Executive Sales Performance Dashboard

Aplikasi dashboard sales interaktif berbasis Streamlit.
Dibuat modular agar mudah dimodifikasi.
"""

import io
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================================================
# 0. KONFIGURASI HALAMAN & KONSTANTA
# =========================================================

st.set_page_config(
    page_title="Sales Display Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

PRIMARY_COLOR = "#1F4E79"
ACCENT_COLOR = "#2E86C1"
BG_CARD = "#F4F6F8"
POSITIVE_COLOR = "#1E8449"
NEGATIVE_COLOR = "#C0392B"

# =========================================================
# STYLE GLOBAL — TIPOGRAFI PRESISI & GLASS HEADER
# =========================================================
# Catatan: blok CSS ini MURNI kosmetik (font-size, spacing, efek glass).
# Tidak menyentuh logika, kalkulasi, atau struktur data apa pun.
# h1 TIDAK di-override secara blanket agar ukuran font Display Mode (TV) —
# yang memang sengaja dibuat besar untuk dibaca dari jarak jauh — tidak ikut
# mengecil (mempertahankan fitur existing).

GLOBAL_TYPOGRAPHY_CSS = """
<style>
/* ---------- Font family modern & konsisten ---------- */
html, body, [class*="css"], .stApp {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Inter, Roboto,
                 Helvetica, Arial, sans-serif !important;
}

/* ---------- Heading sekunder (bukan judul utama h1) ---------- */
h2 { font-size: 1.08rem !important; font-weight: 700 !important; letter-spacing: -0.01em; }
h3 { font-size: 0.98rem !important; font-weight: 650 !important; }
h4 { font-size: 0.90rem !important; font-weight: 650 !important; }

section[data-testid="stSidebar"] h2 {
    font-size: 0.95rem !important;
    margin-top: 0.5rem !important;
    margin-bottom: 0.25rem !important;
}

/* ---------- Label & teks widget sidebar ---------- */
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] .stMarkdown p {
    font-size: 0.80rem !important;
}
section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
    font-size: 0.73rem !important;
    line-height: 1.35 !important;
}

/* ---------- Caption umum ---------- */
[data-testid="stCaptionContainer"] { font-size: 0.78rem !important; }

/* ---------- Button ---------- */
.stButton > button, .stDownloadButton > button {
    font-size: 0.83rem !important;
    font-weight: 600 !important;
    padding: 0.45rem 1rem !important;
    border-radius: 8px !important;
}

/* ---------- Tabs ---------- */
.stTabs [data-baseweb="tab-list"] { gap: 4px; }
.stTabs [data-baseweb="tab"] {
    font-size: 0.83rem !important;
    font-weight: 600 !important;
    padding: 8px 14px !important;
}

/* ---------- Tabel / dataframe ---------- */
[data-testid="stDataFrame"] * { font-size: 0.80rem !important; }

/* ---------- st.metric (dipakai pada tab MTD Sales) ---------- */
[data-testid="stMetricValue"] { font-size: 1.2rem !important; }
[data-testid="stMetricLabel"] { font-size: 0.78rem !important; }

/* ---------- Input / selectbox label ---------- */
.stSelectbox label, .stMultiSelect label, .stDateInput label,
.stTextInput label, .stNumberInput label, .stRadio label {
    font-size: 0.82rem !important;
    font-weight: 600 !important;
}

/* ---------- Expander header ---------- */
[data-testid="stExpander"] summary p { font-size: 0.85rem !important; font-weight: 600 !important; }

/* ---------- GLASS HEADER (morphglass bening) ---------- */
.app-header-glass {
    background: linear-gradient(135deg, rgba(255,255,255,0.70), rgba(228,241,255,0.50));
    backdrop-filter: blur(16px) saturate(180%);
    -webkit-backdrop-filter: blur(16px) saturate(180%);
    border: 1px solid rgba(255,255,255,0.75);
    border-radius: 18px;
    box-shadow: 0 8px 24px rgba(31,78,121,0.10), inset 0 1px 0 rgba(255,255,255,0.55);
    padding: 20px 28px;
    margin-bottom: 18px;
}
</style>
"""


def inject_global_css():
    """Suntik CSS tipografi & glass header. Dipanggil sekali di awal main()."""
    st.markdown(GLOBAL_TYPOGRAPHY_CSS, unsafe_allow_html=True)

# Kandidat nama kolom yang umum dipakai pada data sales retail.
# Digunakan untuk MENYARANKAN mapping kolom, TIDAK untuk mengarang data.
SALES_CANDIDATES = [
    "net sales", "total sales", "sales amount", "sales value", "amount",
    "revenue", "total penjualan", "penjualan", "gross sales", "sales",
]
CATEGORY_CANDIDATES = [
    "category name", "category", "product category", "item category", "kategori",
]
CATEGORY_CODE_CANDIDATES = [
    "category code", "cat code", "kode kategori", "category id", "kode category",
]
STORE_CANDIDATES = [
    "branch name", "store name", "store", "location", "outlet", "cabang", "toko",
]
DATE_CANDIDATES = [
    "date", "tanggal", "transaction date", "order date", "tgl transaksi",
]

# =========================================================
# KONFIGURASI FITUR "GENERATE" (AUTO-DETECT FILE TERBARU)
# =========================================================

# Ekstensi file data yang didukung untuk auto-detection.
VALID_DATA_EXTENSIONS = {".xlsx", ".xls", ".csv"}

# Prioritas folder pencarian (dari yang paling prioritas ke paling akhir).
# Path.home() dipakai agar TIDAK hard-code username Windows — otomatis mengikuti
# home directory user yang sedang login, di OS apa pun.
DOWNLOAD_DIRECTORIES = [
    Path.home() / "Downloads",
    Path.home() / "Download",
    Path.home() / "Desktop",
    Path.home() / "Documents",
]

# Label format file untuk ditampilkan ke user.
FORMAT_LABELS = {
    ".xlsx": "Excel (.xlsx)",
    ".xls": "Excel (.xls)",
    ".csv": "CSV (.csv)",
}

# Prefix/suffix file yang harus diabaikan (file temporary / sedang terkunci).
TEMP_FILE_PREFIXES = ("~$",)
TEMP_FILE_SUFFIXES = (".tmp", ".temp")


# =========================================================
# 1. DATA PROCESSING
# =========================================================

def clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Membersihkan nama kolom dari spasi berlebih dan kolom 'Unnamed'."""
    df = df.copy()
    new_cols = []
    for col in df.columns:
        c = str(col).strip()
        c = re.sub(r"\s+", " ", c)
        new_cols.append(c)
    df.columns = new_cols
    # Buang kolom yang benar-benar kosong (semua NaN) hasil dari 'Unnamed'
    unnamed_empty = [
        c for c in df.columns
        if c.lower().startswith("unnamed") and df[c].isna().all()
    ]
    if unnamed_empty:
        df = df.drop(columns=unnamed_empty)
    return df


def coerce_numeric(series: pd.Series) -> pd.Series:
    """Mengubah kolom menjadi numerik, menangani format ribuan/koma jika perlu."""
    if pd.api.types.is_numeric_dtype(series):
        return series
    cleaned = (
        series.astype(str)
        .str.replace(r"[^\d,.\-]", "", regex=True)
        .str.replace(".", "", regex=False)   # pemisah ribuan gaya ID
        .str.replace(",", ".", regex=False)  # koma desimal -> titik
    )
    return pd.to_numeric(cleaned, errors="coerce")


@st.cache_data(show_spinner=False)
def load_workbook_sheets(file_bytes: bytes, file_name: str):
    """Membaca file Excel/CSV dan mengembalikan dict {nama_sheet: DataFrame}."""
    sheets = {}
    if file_name.lower().endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes))
        sheets["Sheet1"] = clean_column_names(df)
    else:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
        for name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=name)
            if df.empty:
                continue
            sheets[name] = clean_column_names(df)
    return sheets


@st.cache_data(show_spinner=False)
def load_workbook_sheets_from_path(file_path_str: str, mtime: float):
    """Membaca file Excel/CSV dari path di disk (dipakai oleh fitur Generate).

    Parameter `mtime` disertakan sebagai bagian dari cache key agar cache
    otomatis batal (rebuild) ketika file di disk berubah/di-update, tanpa
    perlu membaca ulang file yang isinya tidak berubah."""
    path = Path(file_path_str)
    suffix = path.suffix.lower()
    sheets = {}

    if suffix == ".csv":
        df = None
        last_err = None
        for encoding in ("utf-8", "utf-8-sig", "latin-1"):
            try:
                df = pd.read_csv(path, encoding=encoding)
                break
            except (UnicodeDecodeError, UnicodeError) as e:
                last_err = e
                continue
        if df is None:
            raise ValueError(
                f"Gagal membaca CSV dengan encoding utf-8 / utf-8-sig / latin-1. "
                f"Detail: {last_err}"
            )
        sheets["Sheet1"] = clean_column_names(df)

    elif suffix == ".xlsx":
        xls = pd.ExcelFile(path, engine="openpyxl")
        for name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=name)
            if df.empty:
                continue
            sheets[name] = clean_column_names(df)

    elif suffix == ".xls":
        try:
            xls = pd.ExcelFile(path, engine="xlrd")
        except ImportError as e:
            raise ImportError(
                "Library 'xlrd' dibutuhkan untuk membaca file .xls. "
                "Install dengan: pip install xlrd"
            ) from e
        for name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=name)
            if df.empty:
                continue
            sheets[name] = clean_column_names(df)

    else:
        raise ValueError(f"Format file '{suffix}' tidak didukung.")

    return sheets


def is_temp_file(path: Path) -> bool:
    """Deteksi file temporary/lock (mis. ~$sales.xlsx, file.tmp) yang harus diabaikan."""
    name = path.name
    if name.startswith(TEMP_FILE_PREFIXES):
        return True
    if path.suffix.lower() in TEMP_FILE_SUFFIXES:
        return True
    return False


def scan_directory_for_data_files(directory: Path):
    """Scan non-recursive satu folder untuk file .xlsx/.xls/.csv yang valid.
    Tidak melakukan scanning seluruh hard disk — hanya isi langsung folder ini."""
    if not directory.exists() or not directory.is_dir():
        return []
    candidates = []
    try:
        for entry in directory.iterdir():
            try:
                if not entry.is_file():
                    continue
            except OSError:
                continue
            if entry.suffix.lower() not in VALID_DATA_EXTENSIONS:
                continue
            if is_temp_file(entry):
                continue
            try:
                entry.stat()  # pastikan file dapat diakses (bukan permission denied)
            except (PermissionError, OSError):
                continue
            candidates.append(entry)
    except PermissionError:
        return []
    return candidates


def find_latest_data_file():
    """Cari file data terbaru berdasarkan prioritas folder + modified time.

    Logika:
    1. Iterasi DOWNLOAD_DIRECTORIES sesuai urutan prioritas.
    2. Folder pertama yang punya minimal 1 file valid (.xlsx/.xls/.csv, bukan
       file temporary) dipakai sebagai sumber pencarian.
    3. Dalam folder tersebut, pilih file dengan modified time TERBARU.

    Return: (Path file terpilih, Path folder yang dipakai) atau (None, None)
    jika tidak ada file ditemukan sama sekali di semua folder kandidat.
    """
    for directory in DOWNLOAD_DIRECTORIES:
        candidates = scan_directory_for_data_files(directory)
        if not candidates:
            continue
        candidates_sorted = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
        return candidates_sorted[0], directory
    return None, None


def guess_column(columns, candidates):
    """Menebak kolom yang paling sesuai berdasarkan daftar kandidat (case-insensitive)."""
    lower_map = {c.lower(): c for c in columns}
    for cand in candidates:
        if cand in lower_map:
            return lower_map[cand]
    # partial match
    for cand in candidates:
        for col_lower, original in lower_map.items():
            if cand in col_lower:
                return original
    return None


def pick_best_sheet(sheets: dict):
    """Memilih sheet yang paling mungkin berisi data sales (paling banyak baris & ada kolom numerik)."""
    best_name, best_score = None, -1
    for name, df in sheets.items():
        numeric_cols = df.select_dtypes(include=[np.number]).shape[1]
        score = len(df) * 1 + numeric_cols * 100
        if score > best_score:
            best_name, best_score = name, score
    return best_name


# =========================================================
# 2. FORMAT ANGKA (RUPIAH / INDONESIA)
# =========================================================

def format_number_id(value, decimals: int = 0) -> str:
    """Format angka gaya Indonesia: 1.234.567 (titik sebagai pemisah ribuan)."""
    if pd.isna(value):
        return "-"
    try:
        value = float(value)
    except (ValueError, TypeError):
        return str(value)
    fmt = f"{{:,.{decimals}f}}"
    text = fmt.format(value)
    text = text.replace(",", "#").replace(".", ",").replace("#", ".")
    return text


def format_rupiah(value, decimals: int = 0) -> str:
    return f"Rp {format_number_id(value, decimals)}"


# =========================================================
# 3. SIDEBAR - GENERATE (AUTO-DETECT FILE) & MAPPING KOLOM
# =========================================================

def init_session_state():
    """Inisialisasi state yang dipakai fitur Generate agar persist antar-rerun."""
    if "generated_file" not in st.session_state:
        st.session_state.generated_file = None
    if "sheets_cache" not in st.session_state:
        st.session_state.sheets_cache = None
    if "load_error" not in st.session_state:
        st.session_state.load_error = None
    if "generate_ran_once" not in st.session_state:
        st.session_state.generate_ran_once = False


def _build_file_info(file_path: Path, used_directory: Path):
    mtime = file_path.stat().st_mtime
    suffix = file_path.suffix.lower()
    return {
        "path": str(file_path),
        "name": file_path.name,
        "mtime": mtime,
        "modified_str": datetime.fromtimestamp(mtime).strftime("%d %B %Y %H:%M"),
        "format_label": FORMAT_LABELS.get(suffix, suffix),
        "directory": str(used_directory),
    }


def run_generate_process():
    """Jalankan proses Generate: scan folder → temukan file terbaru → baca → validasi.
    Semua langkah ditampilkan sebagai loading state berurutan."""
    with st.status("⚡ Menjalankan Generate Data...", expanded=True) as status:
        st.write("🔍 Searching latest data...")
        st.write("📂 Checking Downloads folder...")
        latest_file, used_dir = find_latest_data_file()

        if latest_file is None:
            status.update(label="⚠️ Tidak ditemukan file", state="error", expanded=True)
            st.session_state.generated_file = None
            st.session_state.sheets_cache = None
            st.session_state.load_error = {"type": "no_file"}
            st.session_state.generate_ran_once = True
            return

        st.write("📊 Reading latest file...")
        try:
            sheets = load_workbook_sheets_from_path(str(latest_file), latest_file.stat().st_mtime)
        except PermissionError:
            status.update(label="❌ File sedang digunakan", state="error", expanded=True)
            st.session_state.generated_file = None
            st.session_state.sheets_cache = None
            st.session_state.load_error = {
                "type": "locked",
                "file_name": latest_file.name,
                "message": "File sedang digunakan/terkunci oleh aplikasi lain (mis. masih terbuka di Excel).",
            }
            st.session_state.generate_ran_once = True
            return
        except Exception as e:
            status.update(label="❌ Gagal membaca file", state="error", expanded=True)
            st.session_state.generated_file = None
            st.session_state.sheets_cache = None
            st.session_state.load_error = {
                "type": "read_error",
                "file_name": latest_file.name,
                "message": str(e),
            }
            st.session_state.generate_ran_once = True
            return

        st.write("⚙️ Processing data...")
        if not sheets:
            status.update(label="❌ File tidak memiliki data valid", state="error", expanded=True)
            st.session_state.generated_file = None
            st.session_state.sheets_cache = None
            st.session_state.load_error = {
                "type": "empty",
                "file_name": latest_file.name,
                "message": "File tidak memiliki sheet/data yang valid (kosong).",
            }
            st.session_state.generate_ran_once = True
            return

        st.write("✅ Data ready")
        status.update(label="✅ Data ready", state="complete", expanded=False)

    st.session_state.generated_file = _build_file_info(latest_file, used_dir)
    st.session_state.sheets_cache = sheets
    st.session_state.load_error = None
    st.session_state.generate_ran_once = True


def sidebar_generate_section():
    """Panel sidebar untuk fitur Generate — menggantikan file_uploader manual."""
    st.sidebar.markdown("## ⚡ Sumber Data")

    # CSS kecil agar teks tombol sidebar tidak pernah wrap ke baris baru,
    # meski lebar sidebar sempit — teks tetap presisi satu baris.
    st.markdown(
        """
        <style>
        section[data-testid="stSidebar"] .stButton > button {
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            font-size: 0.92rem;
            padding-left: 0.5rem;
            padding-right: 0.5rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.sidebar.caption(
        "Aplikasi otomatis mencari file data (.xlsx / .xls / .csv) TERBARU "
        "di folder Downloads komputer Anda. Tidak perlu pilih file manual."
    )

    generate_clicked = st.sidebar.button(
        "⚡ GENERATE DATA", use_container_width=True, type="primary"
    )
    refresh_clicked = st.sidebar.button(
        "🔄 Refresh / Generate Ulang", use_container_width=True
    )

    if generate_clicked or refresh_clicked:
        run_generate_process()

    info = st.session_state.generated_file
    if info:
        st.sidebar.success("✓ Data berhasil ditemukan")
        st.sidebar.markdown(
            f"""
**File:**
`{info['name']}`

**Lokasi:**
`{info['path']}`

**Last Modified:**
{info['modified_str']}

**Format:**
{info['format_label']}
            """
        )


def render_welcome_screen():
    """Tampilan awal sebelum ada data — Generate hanya berjalan saat tombol diklik."""
    st.markdown(
        f"""
        <div class="app-header-glass" style="text-align:center; padding:56px 24px;">
            <h1 style="color:{PRIMARY_COLOR}; font-size:26px; font-weight:800; margin-bottom:4px;">📊 Sales Display Analyzer</h1>
            <p style="color:#5D6D7E; font-size:14px; margin:0;">Executive Sales Performance Dashboard</p>
            <p style="color:#AEB6BF; font-size:11.5px; margin-top:3px;">Author : Rachmat Hidayat</p>
            <p style="color:#85929E; font-size:13px; margin-top:24px;">
                👈 Klik tombol <b>⚡ GENERATE DATA</b> pada sidebar untuk mencari dan
                menganalisis file data (.xlsx / .xls / .csv) terbaru secara otomatis
                dari folder Downloads Anda.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_generate_error(error_info: dict):
    """Tampilkan pesan error fitur Generate dengan bahasa yang mudah dipahami user."""
    error_type = error_info.get("type")

    if error_type == "no_file":
        searched = "\n".join(f"- `{d}`" for d in DOWNLOAD_DIRECTORIES)
        st.warning(
            "⚠️ **Tidak ditemukan file Excel atau CSV.**\n\n"
            "Silakan pastikan file `.xlsx`, `.xls`, atau `.csv` sudah tersedia "
            "di salah satu folder berikut, lalu klik **Refresh**:\n\n"
            f"{searched}"
        )
        return

    file_name = error_info.get("file_name", "-")
    message = error_info.get("message", "")

    if error_type == "locked":
        st.error(
            "❌ **File sedang digunakan.**\n\n"
            f"**File:**\n{file_name}\n\n"
            f"**Masalah:**\n{message}\n\n"
            "Tutup file tersebut di aplikasi lain (mis. Excel), lalu klik **Refresh**."
        )
    elif error_type == "empty":
        st.error(
            "❌ **Data tidak dapat diproses.**\n\n"
            f"**File:**\n{file_name}\n\n"
            f"**Masalah:**\n{message}"
        )
    else:
        st.error(
            "❌ **Data tidak dapat diproses.**\n\n"
            f"**File:**\n{file_name}\n\n"
            f"**Masalah:**\nFile tidak dapat dibaca (kemungkinan corrupt, format tidak "
            f"valid, atau permission denied).\n\nDetail teknis: {message}"
        )


def render_source_status(info: dict):
    """Info kecil sumber data yang sedang dipakai (ditampilkan di atas dashboard)."""
    st.caption(f"📄 Source: {info['name']}  •  🕒 Updated: {info['modified_str']}")


def sidebar_column_mapping(df: pd.DataFrame):
    st.sidebar.markdown("## 🧭 Pemetaan Kolom")
    st.sidebar.caption("Aplikasi mendeteksi kolom secara otomatis. Sesuaikan bila perlu.")

    columns = list(df.columns)

    sales_guess = guess_column(columns, SALES_CANDIDATES)
    category_guess = guess_column(columns, CATEGORY_CANDIDATES)
    category_code_guess = guess_column(columns, CATEGORY_CODE_CANDIDATES)
    store_guess = guess_column(columns, STORE_CANDIDATES)
    date_guess = guess_column(columns, DATE_CANDIDATES)

    numeric_cols = [c for c in columns if pd.api.types.is_numeric_dtype(df[c])] or columns

    sales_col = st.sidebar.selectbox(
        "Kolom Nilai Sales *",
        options=numeric_cols,
        index=numeric_cols.index(sales_guess) if sales_guess in numeric_cols else 0,
    )
    category_col = st.sidebar.selectbox(
        "Kolom Kategori *",
        options=columns,
        index=columns.index(category_guess) if category_guess in columns else 0,
    )

    category_code_options = ["(Tidak ada)"] + columns
    category_code_default = (
        category_code_options.index(category_code_guess) if category_code_guess in columns else 0
    )
    category_code_sel = st.sidebar.selectbox(
        "Kolom Category Code", options=category_code_options, index=category_code_default,
        help="Opsional. Jika diisi, tab Ranking Category akan menampilkan ranking tambahan berdasarkan kode kategori.",
    )
    category_code_col = None if category_code_sel == "(Tidak ada)" else category_code_sel

    store_options = ["(Tidak ada)"] + columns
    store_default = store_options.index(store_guess) if store_guess in columns else 0
    store_sel = st.sidebar.selectbox("Kolom Store / Location", options=store_options, index=store_default)
    store_col = None if store_sel == "(Tidak ada)" else store_sel

    date_options = ["(Tidak ada)"] + columns
    date_default = date_options.index(date_guess) if date_guess in columns else 0
    date_sel = st.sidebar.selectbox("Kolom Tanggal", options=date_options, index=date_default)
    date_col = None if date_sel == "(Tidak ada)" else date_sel

    return sales_col, category_col, category_code_col, store_col, date_col


# =========================================================
# 4. VALIDASI STRUKTUR DATA
# =========================================================

def validate_structure(df: pd.DataFrame, sales_col: str, category_col: str):
    missing = []
    if sales_col not in df.columns:
        missing.append(sales_col)
    if category_col not in df.columns:
        missing.append(category_col)
    if missing:
        st.error(
            "❌ Struktur file tidak sesuai.\n\n"
            f"**Kolom yang dibutuhkan tapi tidak ditemukan:** {', '.join(missing)}\n\n"
            f"**Kolom yang tersedia pada file:** {', '.join(df.columns)}"
        )
        st.stop()

    if sales_col == category_col:
        st.error(
            "❌ Data tidak dapat diproses.\n\n"
            f"**Masalah:**\nKolom Sales dan Kolom Kategori tidak boleh sama "
            f"(keduanya terdeteksi sebagai `{sales_col}`).\n\n"
            f"**Kolom yang tersedia pada file:** {', '.join(df.columns)}\n\n"
            "Silakan sesuaikan pemetaan kolom pada sidebar."
        )
        st.stop()

    if df.empty:
        st.error("❌ File tidak berisi data (kosong).")
        st.stop()


# =========================================================
# 5. FILTER & AGGREGASI
# =========================================================

def apply_filters(df, sales_col, category_col, store_col, date_col):
    st.sidebar.markdown("## 🔎 Filter")

    filtered = df.copy()

    if store_col:
        stores = sorted(filtered[store_col].dropna().astype(str).unique().tolist())
        sel_stores = st.sidebar.multiselect("Store / Location", options=stores, default=stores)
        if sel_stores:
            filtered = filtered[filtered[store_col].astype(str).isin(sel_stores)]

    categories = sorted(filtered[category_col].dropna().astype(str).unique().tolist())
    sel_categories = st.sidebar.multiselect("Category", options=categories, default=categories)
    if sel_categories:
        filtered = filtered[filtered[category_col].astype(str).isin(sel_categories)]

    if date_col and date_col in filtered.columns:
        parsed_dates = pd.to_datetime(filtered[date_col], errors="coerce")
        valid_dates = parsed_dates.dropna()
        if not valid_dates.empty:
            min_d, max_d = valid_dates.min().date(), valid_dates.max().date()
            date_range = st.sidebar.date_input(
                "Rentang Tanggal", value=(min_d, max_d), min_value=min_d, max_value=max_d
            )
            if isinstance(date_range, tuple) and len(date_range) == 2:
                start, end = date_range
                mask = (parsed_dates.dt.date >= start) & (parsed_dates.dt.date <= end)
                filtered = filtered[mask.fillna(False)]

    top_n_option = st.sidebar.selectbox(
        "Top Ranking Category", options=["Top 5", "Top 10", "Top 15", "Top 20", "All Category"], index=1
    )

    if filtered.empty:
        st.warning("⚠️ Tidak ada data yang sesuai dengan filter yang dipilih. Silakan ubah filter.")
        st.stop()

    return filtered, top_n_option


def top_n_to_int(label: str, max_len: int):
    if label == "All Category":
        return max_len
    return int(label.replace("Top ", ""))


# =========================================================
# 6. KALKULASI RINGKASAN
# =========================================================

def compute_category_summary(df, sales_col, category_col):
    summary = (
        df.groupby(category_col, dropna=False)[sales_col]
        .sum()
        .reset_index()
        .rename(columns={category_col: "Category", sales_col: "Sales"})
        .sort_values("Sales", ascending=False)
        .reset_index(drop=True)
    )
    total_sales = summary["Sales"].sum()
    summary["Rank"] = np.arange(1, len(summary) + 1)
    summary["Contribution %"] = np.where(
        total_sales != 0, summary["Sales"] / total_sales * 100, 0
    )
    return summary[["Rank", "Category", "Sales", "Contribution %"]]


def compute_category_code_summary(df, sales_col, category_col, category_code_col):
    """Ranking sales berdasarkan Category Code (opsional, jika kolom tersedia),
    disertai Category Name yang paling dominan (sales terbesar) untuk kode tsb."""
    if not category_code_col or category_code_col not in df.columns:
        return None

    # Mapping kode -> nama kategori yang paling relevan (sales terbesar) untuk
    # menangani kemungkinan satu kode terhubung ke lebih dari satu nama.
    name_map = (
        df.groupby([category_code_col, category_col], dropna=False)[sales_col]
        .sum()
        .reset_index()
        .sort_values(sales_col, ascending=False)
        .drop_duplicates(subset=[category_code_col])
        .set_index(category_code_col)[category_col]
    )

    summary = (
        df.groupby(category_code_col, dropna=False)[sales_col]
        .sum()
        .reset_index()
        .rename(columns={category_code_col: "Category Code", sales_col: "Sales"})
        .sort_values("Sales", ascending=False)
        .reset_index(drop=True)
    )
    summary["Category Name"] = summary["Category Code"].map(name_map)
    total_sales = summary["Sales"].sum()
    summary["Rank"] = np.arange(1, len(summary) + 1)
    summary["Contribution %"] = np.where(
        total_sales != 0, summary["Sales"] / total_sales * 100, 0
    )
    return summary[["Rank", "Category Code", "Category Name", "Sales", "Contribution %"]]


def compute_mtd_summary(df, sales_col, date_col):
    """Menghitung sales harian dan sales kumulatif (Month-to-Date).
    Kumulatif di-reset otomatis setiap awal bulan kalender baru."""
    if not date_col or date_col not in df.columns:
        return None

    d = df.copy()
    d[date_col] = pd.to_datetime(d[date_col], errors="coerce")
    d = d.dropna(subset=[date_col])
    if d.empty:
        return None

    daily = (
        d.groupby(d[date_col].dt.date)[sales_col]
        .sum()
        .reset_index()
        .rename(columns={date_col: "Date", sales_col: "Daily Sales"})
    )
    daily["Date"] = pd.to_datetime(daily["Date"])
    daily = daily.sort_values("Date").reset_index(drop=True)
    daily["Month"] = daily["Date"].dt.to_period("M").astype(str)
    daily["Cumulative Sales"] = daily.groupby("Month")["Daily Sales"].cumsum()
    return daily


def compute_store_summary(df, sales_col, store_col):
    if not store_col:
        return None
    summary = (
        df.groupby(store_col, dropna=False)[sales_col]
        .agg(Sales="sum", Transactions="count", Average="mean")
        .reset_index()
        .rename(columns={store_col: "Store"})
        .sort_values("Sales", ascending=False)
        .reset_index(drop=True)
    )
    total_sales = summary["Sales"].sum()
    summary["Rank"] = np.arange(1, len(summary) + 1)
    summary["Contribution %"] = np.where(
        total_sales != 0, summary["Sales"] / total_sales * 100, 0
    )
    return summary[["Rank", "Store", "Sales", "Contribution %", "Average"]]


# =========================================================
# 7. KPI CARDS
# =========================================================

def render_kpi_cards(df, sales_col, category_col, store_col, cat_summary):
    total_sales = df[sales_col].sum()
    total_category = df[category_col].nunique()
    total_store = df[store_col].nunique() if store_col else 0
    avg_sales = df[sales_col].mean()
    highest_cat = cat_summary.iloc[0]["Category"] if not cat_summary.empty else "-"
    lowest_cat = cat_summary.iloc[-1]["Category"] if not cat_summary.empty else "-"

    cols = st.columns(3)
    kpis = [
        ("💰 Total Sales", format_rupiah(total_sales)),
        ("🗂️ Total Category", format_number_id(total_category)),
        ("🏬 Total Store / Location", format_number_id(total_store) if store_col else "N/A"),
        ("📊 Average Sales", format_rupiah(avg_sales)),
        ("🏆 Highest Sales Category", highest_cat),
        ("📉 Lowest Sales Category", lowest_cat),
    ]
    for i, (label, value) in enumerate(kpis):
        with cols[i % 3]:
            st.markdown(
                f"""
                <div style="background-color:{BG_CARD}; padding:15px 14px; border-radius:11px;
                            border-left:4px solid {PRIMARY_COLOR}; margin-bottom:14px;">
                    <div style="font-size:11.5px; color:#5D6D7E; font-weight:650; letter-spacing:0.01em;">{label}</div>
                    <div style="font-size:19px; color:{PRIMARY_COLOR}; font-weight:750; margin-top:3px;">
                        {value}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
    return total_sales, total_category, total_store, avg_sales, highest_cat, lowest_cat


# =========================================================
# 8. VISUALISASI (PLOTLY)
# =========================================================

def chart_top_category(cat_summary, n):
    data = cat_summary.head(n).sort_values("Sales", ascending=True)
    fig = go.Figure(
        go.Bar(
            x=data["Sales"],
            y=data["Category"],
            orientation="h",
            marker=dict(color=ACCENT_COLOR),
            text=[format_number_id(v) for v in data["Sales"]],
            textposition="outside",
            hovertemplate="<b>%{y}</b><br>Sales: Rp %{x:,.0f}<extra></extra>",
        )
    )
    fig.update_layout(
        title=f"Top {n} Category Sales",
        xaxis_title="Sales",
        yaxis_title="",
        template="plotly_white",
        height=max(400, 32 * len(data)),
        margin=dict(l=10, r=60, t=50, b=10),
    )
    return fig


def chart_store_ranking(store_summary):
    data = store_summary.sort_values("Sales", ascending=True)
    fig = go.Figure(
        go.Bar(
            x=data["Sales"],
            y=data["Store"],
            orientation="h",
            marker=dict(color=PRIMARY_COLOR),
            text=[format_number_id(v) for v in data["Sales"]],
            textposition="outside",
            hovertemplate="<b>%{y}</b><br>Sales: Rp %{x:,.0f}<extra></extra>",
        )
    )
    fig.update_layout(
        title="Sales by Store / Location",
        xaxis_title="Sales",
        yaxis_title="",
        template="plotly_white",
        height=max(400, 32 * len(data)),
        margin=dict(l=10, r=60, t=50, b=10),
    )
    return fig


def chart_contribution_pie(cat_summary, n=10):
    top = cat_summary.head(n).copy()
    others_sum = cat_summary["Sales"].iloc[n:].sum()
    if others_sum > 0:
        top = pd.concat(
            [top, pd.DataFrame([{"Category": "Others", "Sales": others_sum}])],
            ignore_index=True,
        )
    fig = px.pie(
        top, names="Category", values="Sales", hole=0.45,
        color_discrete_sequence=px.colors.sequential.Blues_r,
    )
    fig.update_traces(textinfo="percent+label", hovertemplate="%{label}<br>Rp %{value:,.0f}<extra></extra>")
    fig.update_layout(title="Sales Contribution by Category", template="plotly_white", height=450)
    return fig


def chart_mtd_sales(data):
    """Bar chart sales harian + line chart sales kumulatif (MTD) dual-axis."""
    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=data["Date"],
            y=data["Daily Sales"],
            name="Daily Sales",
            marker=dict(color=ACCENT_COLOR),
            text=[format_number_id(v) for v in data["Daily Sales"]],
            textposition="outside",
            hovertemplate="%{x|%d %b %Y}<br>Daily Sales: Rp %{y:,.0f}<extra></extra>",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=data["Date"],
            y=data["Cumulative Sales"],
            name="Cumulative Sales (MTD)",
            mode="lines+markers",
            line=dict(color=PRIMARY_COLOR, width=3),
            marker=dict(size=7),
            yaxis="y2",
            hovertemplate="%{x|%d %b %Y}<br>MTD: Rp %{y:,.0f}<extra></extra>",
        )
    )
    fig.update_layout(
        title="Month-to-Date (MTD) Sales",
        xaxis_title="Tanggal",
        yaxis=dict(title="Daily Sales"),
        yaxis2=dict(title="Cumulative Sales (MTD)", overlaying="y", side="right", showgrid=False),
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        height=480,
        margin=dict(l=10, r=10, t=70, b=10),
    )
    return fig


# =========================================================
# 9. EXPORT EXCEL (OPENPYXL) - PRINT-FRIENDLY
# =========================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(size=10, italic=True, color="5D6D7E")
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"), right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"), bottom=Side(style="thin", color="D5D8DC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def autofit_columns(ws, df, start_col=1, sample_size=1000):
    """Menghitung lebar kolom otomatis. Untuk dataset besar, gunakan sampel baris
    agar tetap cepat (lebar kolom tetap representatif tanpa memindai seluruh data)."""
    sample_df = df.sample(n=sample_size, random_state=42) if len(df) > sample_size else df
    for i, col in enumerate(df.columns, start=start_col):
        max_len = max([len(str(col))] + [len(str(v)) for v in sample_df[col].astype(str)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 4, 12), 45)


def write_title_block(ws, title, subtitle, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = LEFT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = SUBTITLE_FONT
    c2.alignment = LEFT
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16


def write_dataframe(ws, df, start_row, number_cols=None, fast_mode=None):
    """Menulis DataFrame ke worksheet dengan header ternormalisasi.

    fast_mode=True melewati styling per-cell (border/alignment) untuk mempercepat
    penulisan data besar. Default otomatis aktif jika baris > 3000 agar export
    tetap responsif untuk dataset besar (mis. Detail Data ribuan baris)."""
    number_cols = number_cols or []
    n_cols = len(df.columns)
    if fast_mode is None:
        fast_mode = len(df) > 3000

    for j, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    if fast_mode:
        # Bulk write menggunakan ws.append — jauh lebih cepat untuk data besar
        # karena menghindari overhead pembuatan objek Border/Alignment per sel.
        for row in df.itertuples(index=False, name=None):
            ws.append(row)
        end_row = start_row + len(df)

        # Format angka diterapkan per-kolom menggunakan integer indexing (ws.cell),
        # jauh lebih cepat dibanding akses via referensi string ("A1"-style).
        format_targets = []
        for col_name in number_cols:
            if col_name in df.columns:
                format_targets.append((list(df.columns).index(col_name) + 1, '#,##0'))
        for col_name in df.columns:
            if "%" in str(col_name):
                format_targets.append((list(df.columns).index(col_name) + 1, '0.00"%"'))

        for col_idx, fmt in format_targets:
            for row_idx in range(start_row + 1, end_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = fmt
    else:
        for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
            for j, col_name in enumerate(row.index, start=1):
                value = row[col_name]
                cell = ws.cell(row=i, column=j, value=value)
                cell.border = THIN_BORDER
                if col_name in number_cols:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif "%" in str(col_name):
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = LEFT
        end_row = start_row + len(df)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(n_cols)}{end_row}"
    return end_row


def setup_print(ws, n_cols, end_row, start_row=1, landscape=True):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{get_column_letter(n_cols)}{end_row}"
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


@st.cache_data(show_spinner="⏳ Menyiapkan file Excel Report...")
def build_excel_report(total_sales, total_category, total_store, avg_sales,
                        cat_summary, store_summary, detail_df, sales_col,
                        period_text):
    wb = Workbook()

    # ---- Sheet 1: Executive Summary ----
    ws1 = wb.active
    ws1.title = "Executive Summary"
    write_title_block(ws1, "SALES PERFORMANCE REPORT", f"Executive Summary — {period_text}", 4)

    summary_rows = [
        ["Metric", "Value"],
        ["Total Sales", total_sales],
        ["Total Category", total_category],
        ["Total Store / Location", total_store],
        ["Average Sales", round(avg_sales, 2)],
        ["Highest Sales Category", cat_summary.iloc[0]["Category"] if not cat_summary.empty else "-"],
        ["Lowest Sales Category", cat_summary.iloc[-1]["Category"] if not cat_summary.empty else "-"],
    ]
    r = 4
    for row in summary_rows:
        for j, val in enumerate(row, start=1):
            cell = ws1.cell(row=r, column=j, value=val)
            cell.border = THIN_BORDER
            if r == 4:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER
            elif j == 2 and isinstance(val, (int, float)) and row[0] in ("Total Sales", "Average Sales"):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
        r += 1

    ws1.cell(row=r + 1, column=1, value="Ranking Category (Top 10)").font = Font(bold=True, size=12, color="1F4E79")
    r += 2
    top10 = cat_summary.head(10).copy()
    top10["Contribution %"] = top10["Contribution %"].round(2)
    end_row1 = write_dataframe(ws1, top10, r, number_cols=["Sales"])
    autofit_columns(ws1, top10)
    setup_print(ws1, 4, end_row1)

    # ---- Sheet 2: Sales by Category ----
    ws2 = wb.create_sheet("Sales by Category")
    write_title_block(ws2, "SALES PERFORMANCE REPORT", f"Sales by Category — {period_text}", 4)
    cat_out = cat_summary.copy()
    cat_out["Contribution %"] = cat_out["Contribution %"].round(2)
    end_row2 = write_dataframe(ws2, cat_out, 4, number_cols=["Sales"])
    autofit_columns(ws2, cat_out)
    setup_print(ws2, 4, end_row2)

    # ---- Sheet 3: Sales by Store ----
    if store_summary is not None:
        ws3 = wb.create_sheet("Sales by Store")
        write_title_block(ws3, "SALES PERFORMANCE REPORT", f"Sales by Store — {period_text}", 5)
        store_out = store_summary.copy()
        store_out["Contribution %"] = store_out["Contribution %"].round(2)
        store_out["Average"] = store_out["Average"].round(0)
        end_row3 = write_dataframe(ws3, store_out, 4, number_cols=["Sales", "Average"])
        autofit_columns(ws3, store_out)
        setup_print(ws3, 5, end_row3)

    # ---- Sheet 4: Detail Data ----
    ws4 = wb.create_sheet("Detail Data")
    write_title_block(ws4, "SALES PERFORMANCE REPORT", f"Detail Data — {period_text}", len(detail_df.columns))
    end_row4 = write_dataframe(ws4, detail_df, 4, number_cols=[sales_col])
    autofit_columns(ws4, detail_df)
    setup_print(ws4, len(detail_df.columns), end_row4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =========================================================
# 10. TAMPILAN UTAMA (DASHBOARD MODE)
# =========================================================

def render_dashboard_mode(df, sales_col, category_col, category_code_col, store_col, date_col):
    validate_structure(df, sales_col, category_col)
    filtered, top_n_label = apply_filters(df, sales_col, category_col, store_col, date_col)

    cat_summary = compute_category_summary(filtered, sales_col, category_col)
    cat_code_summary = compute_category_code_summary(filtered, sales_col, category_col, category_code_col)
    store_summary = compute_store_summary(filtered, sales_col, store_col)
    n = top_n_to_int(top_n_label, len(cat_summary))

    st.markdown(
        f"""
        <div class="app-header-glass" style="padding:16px 24px;">
            <h1 style="color:{PRIMARY_COLOR}; font-size:24px; font-weight:800; margin-bottom:2px;">📊 Sales Display Analyzer</h1>
            <p style="color:#5D6D7E; font-size:13.5px; margin:0;">Executive Sales Performance Dashboard</p>
            <p style="color:#AEB6BF; font-size:11.5px; margin-top:2px;">Author : Rachmat Hidayat</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    render_kpi_cards(filtered, sales_col, category_col, store_col, cat_summary)

    tab_names = [
        "🏆 Top Category", "🏬 Sales by Store", "🥧 Contribution",
        "📋 Ranking Category", "📈 MTD Sales", "📄 Detail Data",
    ]
    tabs = st.tabs(tab_names)

    with tabs[0]:
        st.plotly_chart(chart_top_category(cat_summary, n), use_container_width=True)

    with tabs[1]:
        if store_summary is not None and not store_summary.empty:
            st.plotly_chart(chart_store_ranking(store_summary), use_container_width=True)
            display_store = store_summary.copy()
            display_store["Sales"] = display_store["Sales"].apply(format_rupiah)
            display_store["Average"] = display_store["Average"].apply(format_rupiah)
            display_store["Contribution %"] = display_store["Contribution %"].round(2).astype(str) + "%"
            st.dataframe(display_store, use_container_width=True, hide_index=True)
        else:
            st.info("ℹ️ Data Store/Location tidak tersedia pada file ini.")

    with tabs[2]:
        st.plotly_chart(chart_contribution_pie(cat_summary), use_container_width=True)
        total_contrib = cat_summary["Contribution %"].sum()
        st.caption(f"✅ Total Contribution: {total_contrib:.2f}% (harus mendekati 100%)")

    with tabs[3]:
        st.markdown("#### 📋 Ranking by Category")
        display_cat = cat_summary.copy()
        display_cat["Sales"] = display_cat["Sales"].apply(format_rupiah)
        display_cat["Contribution %"] = display_cat["Contribution %"].round(2).astype(str) + "%"
        st.dataframe(display_cat, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("#### 🔖 Top by Category Code")
        if not category_code_col:
            st.info(
                "ℹ️ Kolom Category Code belum dipilih/terdeteksi. Silakan pilih pada "
                "sidebar bagian **Pemetaan Kolom** (opsional) untuk melihat ranking "
                "berdasarkan kode kategori."
            )
        elif cat_code_summary is None or cat_code_summary.empty:
            st.warning("⚠️ Tidak ada data Category Code yang valid pada data terfilter.")
        else:
            display_cat_code = cat_code_summary.copy()
            display_cat_code["Sales"] = display_cat_code["Sales"].apply(format_rupiah)
            display_cat_code["Contribution %"] = display_cat_code["Contribution %"].round(2).astype(str) + "%"
            st.dataframe(display_cat_code, use_container_width=True, hide_index=True)
            st.caption(
                f"Top Category Code: **{cat_code_summary.iloc[0]['Category Code']}** "
                f"({cat_code_summary.iloc[0]['Category Name']}) "
                f"dengan sales {format_rupiah(cat_code_summary.iloc[0]['Sales'])}"
            )

    with tabs[4]:
        if not date_col:
            st.info(
                "ℹ️ Kolom Tanggal belum dipilih/terdeteksi, sehingga MTD Sales "
                "tidak dapat dihitung. Silakan pilih kolom tanggal pada sidebar "
                "bagian **Pemetaan Kolom**."
            )
        else:
            mtd_data = compute_mtd_summary(filtered, sales_col, date_col)
            if mtd_data is None or mtd_data.empty:
                st.warning("⚠️ Tidak ditemukan data tanggal yang valid untuk menghitung MTD Sales.")
            else:
                months_available = sorted(mtd_data["Month"].unique(), reverse=True)
                selected_month = st.selectbox(
                    "Pilih Bulan (Month-to-Date)", options=months_available, index=0
                )
                month_data = mtd_data[mtd_data["Month"] == selected_month].reset_index(drop=True)

                if month_data.empty:
                    st.warning("⚠️ Tidak ada data pada bulan yang dipilih.")
                else:
                    total_mtd = month_data["Cumulative Sales"].iloc[-1]
                    n_days = len(month_data)
                    avg_daily = month_data["Daily Sales"].mean()
                    last_date = month_data["Date"].max().strftime("%d %B %Y")

                    m1, m2, m3 = st.columns(3)
                    m1.metric("Total MTD Sales", format_rupiah(total_mtd))
                    m2.metric("Jumlah Hari Transaksi", format_number_id(n_days))
                    m3.metric("Rata-rata Sales Harian", format_rupiah(avg_daily))
                    st.caption(f"Sales terkumulasi (Month-to-Date) sampai dengan **{last_date}**.")

                    st.plotly_chart(chart_mtd_sales(month_data), use_container_width=True)

                    display_mtd = month_data.copy()
                    display_mtd["Date"] = display_mtd["Date"].dt.strftime("%d-%m-%Y")
                    display_mtd["Daily Sales"] = display_mtd["Daily Sales"].apply(format_rupiah)
                    display_mtd["Cumulative Sales"] = display_mtd["Cumulative Sales"].apply(format_rupiah)
                    st.dataframe(
                        display_mtd[["Date", "Daily Sales", "Cumulative Sales"]],
                        use_container_width=True, hide_index=True,
                    )

    with tabs[5]:
        st.markdown(f"**Total Baris:** {format_number_id(len(filtered))}")
        search_term = st.text_input("🔍 Cari data (semua kolom)", "")
        show_df = filtered.copy()
        if search_term:
            mask = show_df.astype(str).apply(
                lambda col: col.str.contains(search_term, case=False, na=False)
            ).any(axis=1)
            show_df = show_df[mask]
        st.dataframe(show_df, use_container_width=True, height=420)
        st.caption(f"Menampilkan {format_number_id(len(show_df))} dari {format_number_id(len(filtered))} baris")

    st.markdown("---")
    st.markdown("### ⬇️ Export Report")

    if date_col and date_col in filtered.columns:
        parsed = pd.to_datetime(filtered[date_col], errors="coerce").dropna()
        period_text = f"{parsed.min().date()} s/d {parsed.max().date()}" if not parsed.empty else "N/A"
    else:
        period_text = datetime.now().strftime("%Y-%m-%d")

    excel_buffer = build_excel_report(
        total_sales=filtered[sales_col].sum(),
        total_category=filtered[category_col].nunique(),
        total_store=filtered[store_col].nunique() if store_col else 0,
        avg_sales=filtered[sales_col].mean(),
        cat_summary=cat_summary,
        store_summary=store_summary,
        detail_df=filtered,
        sales_col=sales_col,
        period_text=period_text,
    )

    st.download_button(
        label="📥 Download Excel Report",
        data=excel_buffer,
        file_name=f"Sales_Performance_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================================================
# 11. DISPLAY SALES MODE (UNTUK MONITOR/TV)
# =========================================================

def render_display_mode(df, sales_col, category_col, store_col, date_col):
    st.markdown(
        f"""
        <style>
        .big-kpi-label {{ font-size:20px; color:#EAECEE; font-weight:600; }}
        .big-kpi-value {{ font-size:42px; color:#FFFFFF; font-weight:800; margin-top:4px;}}
        .display-bg {{ background-color:{PRIMARY_COLOR}; padding:22px 26px; border-radius:16px; }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    validate_structure(df, sales_col, category_col)
    filtered = df.copy()
    if store_col:
        filtered = filtered[filtered[store_col].notna()]

    cat_summary = compute_category_summary(filtered, sales_col, category_col)
    store_summary = compute_store_summary(filtered, sales_col, store_col)

    total_sales = filtered[sales_col].sum()
    total_category = filtered[category_col].nunique()
    total_store = filtered[store_col].nunique() if store_col else 0

    st.markdown(
        f"""
        <div class="app-header-glass" style="text-align:center; padding:22px 20px;">
            <h1 style="color:{PRIMARY_COLOR}; font-size:44px; margin-bottom:0;">SALES PERFORMANCE REPORT</h1>
            <p style="color:#5D6D7E; font-size:18px;">{datetime.now().strftime('%A, %d %B %Y — %H:%M')}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns(3)
    for col, label, value in zip(
        [c1, c2, c3],
        ["TOTAL SALES", "TOTAL CATEGORY", "TOTAL STORE"],
        [format_rupiah(total_sales), format_number_id(total_category), format_number_id(total_store)],
    ):
        with col:
            st.markdown(
                f"""<div class="display-bg" style="text-align:center;">
                        <div class="big-kpi-label">{label}</div>
                        <div class="big-kpi-value">{value}</div>
                    </div>""",
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)
    col_a, col_b = st.columns(2)
    with col_a:
        fig = chart_top_category(cat_summary, min(10, len(cat_summary)))
        fig.update_layout(height=520, font=dict(size=14))
        st.plotly_chart(fig, use_container_width=True)
    with col_b:
        if store_summary is not None and not store_summary.empty:
            fig2 = chart_store_ranking(store_summary.head(10))
            fig2.update_layout(height=520, font=dict(size=14))
            st.plotly_chart(fig2, use_container_width=True)
        else:
            fig2 = chart_contribution_pie(cat_summary)
            fig2.update_layout(height=520, font=dict(size=14))
            st.plotly_chart(fig2, use_container_width=True)

    auto_refresh = st.checkbox("🔄 Auto-refresh (30 detik)", value=False, key="display_auto_refresh")
    if auto_refresh:
        st.caption("Dashboard akan menyegarkan otomatis setiap 30 detik.")
        import time
        time.sleep(30)
        st.rerun()


# =========================================================
# 12. MAIN APP
# =========================================================

def main():
    inject_global_css()
    init_session_state()
    sidebar_generate_section()

    file_info = st.session_state.generated_file
    sheets = st.session_state.sheets_cache
    load_error = st.session_state.load_error

    if file_info is None or sheets is None:
        if load_error:
            render_generate_error(load_error)
        else:
            render_welcome_screen()
        return

    if not sheets:
        st.error("❌ File tidak memiliki sheet dengan data yang valid.")
        st.stop()

    sheet_names = list(sheets.keys())
    default_sheet = pick_best_sheet(sheets)

    if len(sheet_names) > 1:
        st.sidebar.markdown("## 📑 Sheet")
        selected_sheet = st.sidebar.selectbox(
            "Pilih Sheet Data Sales", options=sheet_names,
            index=sheet_names.index(default_sheet) if default_sheet in sheet_names else 0,
        )
    else:
        selected_sheet = sheet_names[0]

    df = sheets[selected_sheet]

    with st.sidebar.expander("🔍 Info Struktur File", expanded=False):
        st.write(f"**Sheet aktif:** {selected_sheet}")
        st.write(f"**Jumlah baris:** {format_number_id(len(df))}")
        st.write(f"**Jumlah kolom:** {len(df.columns)}")
        st.write("**Kolom:**")
        st.write(", ".join(df.columns))

    sales_col, category_col, category_code_col, store_col, date_col = sidebar_column_mapping(df)
    validate_structure(df, sales_col, category_col)

    # Pastikan kolom sales numerik (tanpa mengubah data sumber di luar salinan kerja)
    df = df.copy()
    df[sales_col] = coerce_numeric(df[sales_col])
    df = df.dropna(subset=[sales_col])
    df[category_col] = df[category_col].fillna("(Tidak Berkategori)")
    if category_code_col:
        df[category_code_col] = df[category_code_col].fillna("(Tidak Berkode)")

    render_source_status(file_info)

    st.sidebar.markdown("## 🖥️ Mode Tampilan")
    mode = st.sidebar.radio("Pilih Mode", ["Dashboard", "Display Sales (TV/Monitor)"], index=0)

    if mode == "Dashboard":
        render_dashboard_mode(df, sales_col, category_col, category_code_col, store_col, date_col)
    else:
        render_display_mode(df, sales_col, category_col, store_col, date_col)


if __name__ == "__main__":
    main()
